import os
import sys
import time
import pandas as pd
import datetime as dt
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import io

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# FORMATTING CONSTANTS
# ─────────────────────────────────────────────
HEADER_FILL_TRADE   = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_CHECK   = PatternFill("solid", fgColor="375623")
HEADER_FILL_ALLOC   = PatternFill("solid", fgColor="7B2C2C")
HEADER_FILL_MAPPING = PatternFill("solid", fgColor="7B5B00")
HEADER_FONT         = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT           = Font(name="Arial", size=10)
ALT_FILL_LIGHT      = PatternFill("solid", fgColor="EBF3FB")
ALT_FILL_NONE       = PatternFill("solid", fgColor="FFFFFF")
OK_FILL             = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL           = PatternFill("solid", fgColor="FFF2CC")
SUGGEST_FILL        = PatternFill("solid", fgColor="FFF9C4")
THIN                = Side(style='thin', color="D9D9D9")
BORDER              = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─────────────────────────────────────────────
# UI HELPERS
# ─────────────────────────────────────────────
def show_msg(title, message):
    root = tk.Tk(); root.withdraw(); root.lift(); root.attributes('-topmost', True)
    messagebox.showinfo(title, message, parent=root); root.destroy()

def prompt_file(title, filetypes=None):
    root = tk.Tk(); root.withdraw(); root.lift(); root.attributes('-topmost', True)
    path = filedialog.askopenfilename(title=title, filetypes=filetypes or [])
    root.destroy(); return path

def prompt_folder(title):
    root = tk.Tk(); root.withdraw(); root.lift(); root.attributes('-topmost', True)
    path = filedialog.askdirectory(title=title); root.destroy(); return path

def ask_yes_no(title, message):
    root = tk.Tk(); root.withdraw(); root.lift(); root.attributes('-topmost', True)
    result = messagebox.askyesno(title, message, parent=root); root.destroy()
    return result

# ─────────────────────────────────────────────
# FILE LOADING
# ─────────────────────────────────────────────
def convert_html_xls_to_xlsx(input_path):
    with open(input_path, 'rb') as f:
        content = f.read()
    soup = BeautifulSoup(content, 'html.parser')
    tables = pd.read_html(io.StringIO(str(soup)), header=None)
    if not tables:
        raise ValueError("No HTML tables found.")
    raw_df = tables[0]
    raw_df.columns = raw_df.iloc[0].astype(str).str.strip()
    df = raw_df.iloc[1:].reset_index(drop=True)
    output_path = os.path.splitext(input_path)[0] + ".xlsx"
    df.to_excel(output_path, index=False)
    return output_path

def load_excel_safe(path):
    ext = os.path.splitext(path)[-1].lower()
    try:
        if ext == '.xls':
            try:
                return pd.read_excel(path, sheet_name=0, engine='xlrd')
            except Exception:
                return pd.read_excel(convert_html_xls_to_xlsx(path),
                                     sheet_name=0, engine='openpyxl')
        elif ext in ['.xlsx', '.xlsm']:
            return pd.read_excel(path, sheet_name=0, engine='openpyxl')
        else:
            raise ValueError(f"Unsupported extension: {ext}")
    except Exception as e:
        print(f"Failed to read: {path}\n{e}"); sys.exit()

def clean_df(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ─────────────────────────────────────────────
# ALLOCATION HELPERS
# ─────────────────────────────────────────────
def load_allocation(xls, sheet):
    """
    Load allocation sheet. Handles:
    - 'weights' or 'weights (%)' column name
    - Decimal (0.20) or % integer (20) format — normalises to decimal
    - Zero-pads mgr_code and fund_code to 3 digits
    """
    try:
        df = pd.read_excel(xls, sheet_name=sheet).rename(
            columns=lambda x: str(x).strip().lower())
    except Exception:
        return None
    if 'mgr_code' not in df.columns or 'fund_code' not in df.columns:
        return None
    # Accept 'weights (%)' as 'weights'
    if 'weights' not in df.columns and 'weights (%)' in df.columns:
        df.rename(columns={'weights (%)': 'weights'}, inplace=True)
    if 'weights' not in df.columns:
        return None
    df['mgr_code'] = df['mgr_code'].apply(
        lambda x: str(int(float(x))).zfill(3)
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
    df['fund_code'] = df['fund_code'].apply(
        lambda x: str(int(float(x))).zfill(3)
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
    df['weights'] = pd.to_numeric(df['weights'], errors='coerce').fillna(0)
    # Normalise % integers to decimals
    if df['weights'].max() > 1:
        df['weights'] = df['weights'] / 100
    df['key'] = df['mgr_code'] + df['fund_code']
    return df

def prepare_smart_tab(xls, sheet):
    """Display-ready allocation df: weights as %, fund_code zero-padded."""
    df = pd.read_excel(xls, sheet_name=sheet, dtype={'fund_code': str})
    df.columns = [c.strip() for c in df.columns]
    if 'fund_code' in df.columns:
        df['fund_code'] = df['fund_code'].apply(
            lambda x: str(int(float(x))).zfill(3) if pd.notna(x) else x)
    weight_col = 'weights (%)' if 'weights (%)' in df.columns else 'weights'
    if weight_col in df.columns:
        df[weight_col] = pd.to_numeric(df[weight_col], errors='coerce')
        if df[weight_col].dropna().le(1).all():
            df[weight_col] = (df[weight_col] * 100).round(2)
        if weight_col == 'weights':
            df.rename(columns={'weights': 'weights (%)'}, inplace=True)
    return df

# ─────────────────────────────────────────────
# FUND NAME MAP
# ─────────────────────────────────────────────
def build_fund_name_map(xls_list):
    fmap = {}
    for xls in xls_list:
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet).rename(
                    columns=lambda x: str(x).strip().lower())
                if 'mgr_code' not in df.columns or 'fund_code' not in df.columns:
                    continue
                for _, r in df.iterrows():
                    try:
                        mgr  = str(int(float(r['mgr_code']))).zfill(3)
                        fund = str(int(float(r['fund_code']))).zfill(3)
                        name = str(r.get('fund name', '')) \
                               if pd.notna(r.get('fund name', '')) else ''
                        if name:
                            fmap[(mgr, fund)] = name
                    except (ValueError, TypeError):
                        continue
            except Exception:
                continue
    return fmap

# ─────────────────────────────────────────────
# DELTA CALCULATION
# ─────────────────────────────────────────────
def compute_deltas(new_xls, old_xls, port_type):
    """
    Returns merged df with columns:
      mgr_code, fund_code, key, old_w, new_w, delta, fund_name
    Only returns rows where delta != 0.
    """
    new_alloc = load_allocation(new_xls, port_type)
    old_alloc = load_allocation(old_xls, port_type)
    if new_alloc is None or old_alloc is None:
        return None

    new_cols = new_alloc[['key', 'mgr_code', 'fund_code', 'weights']].rename(
        columns={'weights': 'new_w', 'mgr_code': 'mgr_new', 'fund_code': 'fund_new'})
    old_cols = old_alloc[['key', 'mgr_code', 'fund_code', 'weights']].rename(
        columns={'weights': 'old_w', 'mgr_code': 'mgr_old', 'fund_code': 'fund_old'})

    merged = pd.merge(new_cols, old_cols, on='key', how='outer')
    merged['new_w']     = pd.to_numeric(merged['new_w'], errors='coerce').fillna(0)
    merged['old_w']     = pd.to_numeric(merged['old_w'], errors='coerce').fillna(0)
    merged['mgr_code']  = merged['mgr_new'].fillna(merged['mgr_old'])
    merged['fund_code'] = merged['fund_new'].fillna(merged['fund_old'])
    merged['delta']     = (merged['new_w'] - merged['old_w']).round(6)

    # Attach fund names from both files
    name_map = build_fund_name_map([new_xls, old_xls])
    merged['fund_name'] = merged.apply(
        lambda r: name_map.get((r['mgr_code'], r['fund_code']), ''), axis=1)

    return merged[merged['delta'] != 0].copy()

# ─────────────────────────────────────────────
# AUTO-PAIRING LOGIC
# ─────────────────────────────────────────────
def auto_pair(deltas_df):
    """
    Greedy delta matching:
    - Sort switch-out funds by abs(delta) descending
    - Sort switch-in funds by delta descending
    - Match each switch-out fund to switch-in fund(s) whose deltas sum closest
    - If one switch-out delta matches multiple switch-in deltas, split across them
    Returns list of dicts: {MgrCd_OUT, FundCd_OUT, FundName_OUT,
                            MgrCd_IN, FundCd_IN, FundName_IN, delta_out, delta_in}
    """
    out_funds = deltas_df[deltas_df['delta'] < 0].copy()
    in_funds  = deltas_df[deltas_df['delta'] > 0].copy()

    out_funds = out_funds.sort_values('delta').reset_index(drop=True)   # most negative first
    in_funds  = in_funds.sort_values('delta', ascending=False).reset_index(drop=True)

    # remaining capacity for each switch-in fund
    in_remaining = in_funds['delta'].tolist()
    in_list      = in_funds.to_dict('records')

    pairs = []
    for _, out_row in out_funds.iterrows():
        needed = abs(out_row['delta'])
        j = 0
        while needed > 1e-8 and j < len(in_list):
            available = in_remaining[j]
            if available <= 1e-8:
                j += 1
                continue
            take = min(needed, available)
            pairs.append({
                'MgrCd_OUT':   out_row['mgr_code'],
                'FundCd_OUT':  out_row['fund_code'],
                'FundName_OUT': out_row['fund_name'],
                'MgrCd_IN':    in_list[j]['mgr_code'],
                'FundCd_IN':   in_list[j]['fund_code'],
                'FundName_IN': in_list[j]['fund_name'],
                'delta_out':   round(out_row['delta'], 4),
                'delta_in':    round(in_list[j]['delta'], 4),
                'delta_matched': round(take, 4),
            })
            in_remaining[j] -= take
            needed           -= take
            j += 1

    return pairs

# ─────────────────────────────────────────────
# MAPPING FILE WRITER
# ─────────────────────────────────────────────
def write_mapping_suggestion(path, new_xls, old_xls, port_types):
    """
    Write a suggested mapping Excel file with one sheet per portfolio type.
    Rows are pre-filled by auto_pair(). User edits and saves before continuing.
    """
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        for pt in port_types:
            deltas = compute_deltas(new_xls, old_xls, pt)
            if deltas is None or deltas.empty:
                continue

            pairs = auto_pair(deltas)
            if not pairs:
                continue

            df = pd.DataFrame(pairs)[[
                'MgrCd_OUT', 'FundCd_OUT', 'FundName_OUT',
                'MgrCd_IN',  'FundCd_IN',  'FundName_IN',
                'delta_out', 'delta_in', 'delta_matched'
            ]]
            df.insert(0, 'port_type', pt)
            df.to_excel(writer, sheet_name=pt, index=False)

    # Format with openpyxl
    wb = load_workbook(path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # Header row
        for cell in ws[1]:
            cell.fill      = HEADER_FILL_MAPPING
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=True)
            cell.border    = BORDER
        ws.row_dimensions[1].height = 30
        # Data rows + highlight suggestion in yellow
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = ALT_FILL_LIGHT if row_idx % 2 == 0 else ALT_FILL_NONE
            for cell in row:
                cell.font   = BODY_FONT
                cell.border = BORDER
                cell.fill   = fill
        # Note in first row above header
        ws.insert_rows(1)
        note = ws.cell(row=1, column=1)
        note.value = (
            "AUTO-SUGGESTED MAPPING — Please verify each row. "
            "Add/remove/edit rows as needed. "
            "Save this file before clicking OK in the script."
        )
        note.font      = Font(bold=True, italic=True, name="Arial",
                              size=10, color="7B5B00")
        note.alignment = Alignment(horizontal='left')
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=10)
        ws.freeze_panes = 'A3'
        # Column widths
        col_widths = {'A': 14, 'B': 12, 'C': 12, 'D': 45, 'E': 12,
                      'F': 12, 'G': 45, 'H': 12, 'I': 12, 'J': 14}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
    wb.save(path)
    print(f"  Mapping file saved: {path}")

# ─────────────────────────────────────────────
# LOAD CONFIRMED MAPPING
# ─────────────────────────────────────────────
def load_mapping(path):
    """
    Read the confirmed mapping file (all sheets combined).
    Returns DataFrame with columns:
      port_type, MgrCd_OUT, FundCd_OUT, MgrCd_IN, FundCd_IN
    """
    xls  = pd.ExcelFile(path)
    rows = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet)
        df.columns = [str(c).strip() for c in df.columns]
        # Skip note rows (non-data rows where port_type is blank or is the note text)
        if 'port_type' not in df.columns:
            continue
        df = df[df['port_type'].notna()].copy()
        df = df[~df['port_type'].astype(str).str.startswith('AUTO')]
        # Zero-pad fund codes
        for col in ['MgrCd_OUT', 'FundCd_OUT', 'MgrCd_IN', 'FundCd_IN']:
            if col in df.columns:
                df[col] = df[col].apply(
                    lambda x: str(int(float(x))).zfill(3)
                    if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
        rows.append(df[['port_type', 'MgrCd_OUT', 'FundCd_OUT',
                         'MgrCd_IN', 'FundCd_IN']])
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

# ─────────────────────────────────────────────
# CORE TRADE CALCULATION
# ─────────────────────────────────────────────
def compute_switch_trades(rebal_df, new_xls, old_xls, mapping_df, fund_name_map):
    """
    For each account × source fund × destination fund (per mapping):
      - Look up existing qty from rebal input
      - Calculate units out: full qty if new_w=0, else (old-new)/old × qty
      - If source fund maps to multiple destinations, split units by relative delta_in
      - Populate all check columns (N–X) inline
    Returns enriched DataFrame.
    """
    rebal_df = rebal_df.copy()
    rebal_df.columns = [c.strip() for c in rebal_df.columns]
    rebal_df['svc_acct']  = rebal_df['svc_acct'].astype(str).str.strip().str.zfill(7)
    rebal_df['port_type'] = rebal_df['port_type'].astype(str).str.strip().str.replace(
        r'^T_', '', regex=True)
    rebal_df['code'] = rebal_df['code'].apply(
        lambda x: str(int(float(x))).zfill(6)
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
    rebal_df['MgrCd']    = rebal_df['code'].str[:3]
    rebal_df['FundCd']   = rebal_df['code'].str[-3:]
    rebal_df['qty']      = pd.to_numeric(rebal_df['qty'],      errors='coerce').fillna(0)
    rebal_df['mkt_val']  = pd.to_numeric(rebal_df['mkt_val'],  errors='coerce').fillna(0)
    rebal_df['total_pf'] = pd.to_numeric(rebal_df['total_pf'], errors='coerce').fillna(0)

    # Build rebal lookup: (acct, mgr, fund) -> row
    rebal_lookup = {}
    for _, r in rebal_df.iterrows():
        rebal_lookup[(r['svc_acct'], r['MgrCd'], r['FundCd'])] = r

    # Build new alloc weight lookup: (port_type, mgr, fund) -> decimal weight
    new_alloc_lookup = {}
    for sheet in new_xls.sheet_names:
        alloc = load_allocation(new_xls, sheet)
        if alloc is None:
            continue
        for _, r in alloc.iterrows():
            new_alloc_lookup[(sheet, r['mgr_code'], r['fund_code'])] = r['weights']

    # Build old alloc weight lookup: (port_type, mgr, fund) -> decimal weight
    old_alloc_lookup = {}
    for sheet in old_xls.sheet_names:
        alloc = load_allocation(old_xls, sheet)
        if alloc is None:
            continue
        for _, r in alloc.iterrows():
            old_alloc_lookup[(sheet, r['mgr_code'], r['fund_code'])] = r['weights']

    today_str  = dt.datetime.today().strftime("%Y%m%d")
    trade_rows = []

    # Group mapping by (port_type, MgrCd_OUT, FundCd_OUT) to find all destinations
    mapping_grouped = mapping_df.groupby(
        ['port_type', 'MgrCd_OUT', 'FundCd_OUT'])

    for (port_type, mgr_out, fund_out), dest_rows in mapping_grouped:
        # Get all destination funds for this source
        destinations = dest_rows[['MgrCd_IN', 'FundCd_IN']].values.tolist()

        # Get switch-in fund deltas for proportional splitting
        in_deltas = []
        for mgr_in, fund_in in destinations:
            new_w_in = new_alloc_lookup.get((port_type, mgr_in, fund_in), 0)
            old_w_in = old_alloc_lookup.get((port_type, mgr_in, fund_in), 0)
            in_deltas.append(max(new_w_in - old_w_in, 0))
        total_in_delta = sum(in_deltas)

        # Weights for old alloc
        old_w = old_alloc_lookup.get((port_type, mgr_out, fund_out), 0)
        new_w = new_alloc_lookup.get((port_type, mgr_out, fund_out), 0)

        fdsrc  = 'SRS-IA' if port_type.endswith('_SRS') else 'Cash'
        divopt = 'Withdraw' if port_type.startswith('MD') else 'Reinvest'

        # Get all unique accounts holding this port_type
        acct_subset = rebal_df[rebal_df['port_type'] == port_type]
        unique_accts = acct_subset['svc_acct'].unique()

        for acct_no in unique_accts:
            rebal_row = rebal_lookup.get((acct_no, mgr_out, fund_out))
            if rebal_row is None:
                continue
            qty      = rebal_row['qty']
            mkt_val  = rebal_row['mkt_val']
            total_pf = rebal_row['total_pf']
            if qty <= 0:
                continue

            # Units switched out of source fund
            if new_w == 0:
                units_out = qty                                   # full switch
            elif old_w > 0:
                units_out = round(((old_w - new_w) / old_w) * qty, 3)
            else:
                continue
            if units_out <= 0:
                continue

            # Check columns — same for all destination rows of this source fund
            p_val = round(qty, 3)
            q_val = round(units_out, 3)
            r_val = round(qty - units_out, 3)
            s_val = round(mkt_val, 2)
            t_val = round(total_pf, 2)
            u_val = round(units_out / qty, 6)       if qty > 0       else None
            v_val = round(((qty - units_out) / qty * mkt_val) / total_pf, 6) \
                    if (qty > 0 and total_pf > 0)   else None
            w_val = round(new_w * 100, 2)           # target in %
            x_val = port_type

            so_name = fund_name_map.get((mgr_out, fund_out), '')

            # Split units across destinations proportionally by delta_in
            n_dest    = len(destinations)
            allocated = 0.0
            for i, ((mgr_in, fund_in), delta_in) in enumerate(
                    zip(destinations, in_deltas)):
                is_last = (i == n_dest - 1)
                if is_last:
                    units_this = round(units_out - allocated, 3)
                elif total_in_delta > 0:
                    units_this  = round(units_out * (delta_in / total_in_delta), 3)
                    allocated  += units_this
                else:
                    # Equal split if no delta info
                    units_this  = round(units_out / n_dest, 3)
                    allocated  += units_this

                si_name = fund_name_map.get((mgr_in, fund_in), '')

                trade_rows.append({
                    # ── Trades columns A–M ──
                    'AccountNo':   acct_no,
                    'OrderType':   'ESO',
                    'MgrCd':       mgr_out,
                    'FundCd':      fund_out,
                    'FdSrc':       fdsrc,
                    'SIMgrCd':     mgr_in,
                    'SIFundCd':    fund_in,
                    'NetSalesChg': 0,
                    'SwChg':       0,
                    'DivOpt':      divopt,
                    'Units':       units_this,
                    'FACode':      '',
                    'TransDate':   today_str,
                    # ── Check columns N–X ──
                    'Switch Out Fund Name':                       so_name,
                    'Switch In Fund Name':                        si_name,
                    'Units before Rebalance (Switch Out Fund)':   p_val,
                    'Total Units Switched Out':                   q_val,
                    'Units Remaining after Rebalance':            r_val,
                    'MV of Fund before Rebalance (SGD)':          s_val,
                    'Account Portfolio Amount (SGD)':             t_val,
                    'Total Switch Out %':                         u_val,
                    'Switch Out Fund Weight (%) after Rebalance': v_val,
                    'Target Allocation of Switch Out Fund (%)':   w_val,
                    'Portfolio Type':                             x_val,
                })

    return pd.DataFrame(trade_rows) if trade_rows else pd.DataFrame()

# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────
def fmt_sheet(ws, header_fill, freeze='A2', header_row=1, data_start_row=2):
    for cell in ws[header_row]:
        cell.fill      = header_fill
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[header_row].height = 30
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row),
                                  start=data_start_row):
        fill = ALT_FILL_LIGHT if row_idx % 2 == 0 else ALT_FILL_NONE
        for cell in row:
            cell.font   = BODY_FONT
            cell.border = BORDER
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = fill
    if freeze:
        ws.freeze_panes = freeze
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0)
                      for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(
            max(max_len + 2, 10), 50)

def apply_vw_conditional(ws, v_col_idx, w_col_idx, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        v_cell = row[v_col_idx - 1]
        w_cell = row[w_col_idx - 1]
        if v_cell.value is not None and w_cell.value is not None:
            try:
                v_pct = float(v_cell.value) * 100
                w_pct = float(w_cell.value)
                v_cell.fill = OK_FILL if abs(v_pct - w_pct) < 0.5 else WARN_FILL
            except (TypeError, ValueError):
                pass

# ─────────────────────────────────────────────
# OUTPUT FILE WRITER
# ─────────────────────────────────────────────
def write_switch_file(path, trades_df, new_xls, relevant_sheets):
    TRADE_COLS = ['AccountNo', 'OrderType', 'MgrCd', 'FundCd', 'FdSrc',
                  'SIMgrCd', 'SIFundCd', 'NetSalesChg', 'SwChg', 'DivOpt',
                  'Units', 'FACode', 'TransDate']
    CHECK_COLS = TRADE_COLS + [
        'Switch Out Fund Name',
        'Switch In Fund Name',
        'Units before Rebalance (Switch Out Fund)',
        'Total Units Switched Out',
        'Units Remaining after Rebalance',
        'MV of Fund before Rebalance (SGD)',
        'Account Portfolio Amount (SGD)',
        'Total Switch Out %',
        'Switch Out Fund Weight (%) after Rebalance',
        'Target Allocation of Switch Out Fund (%)',
        'Portfolio Type',
    ]

    trades_out = trades_df[TRADE_COLS].copy()
    check_out  = trades_df[CHECK_COLS].copy()

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Sheet 1: Trades — clean values
        trades_out.to_excel(writer, sheet_name='Trades', index=False)

        # Sheet 2: Check — note row 1, headers row 2, data from row 3
        ws_chk = writer.book.create_sheet('Check')
        writer.sheets['Check'] = ws_chk
        ws_chk.cell(row=1, column=1, value=(
            "CHECK TAB — Green = post-switch weight within 0.5% of target.  "
            "Yellow = deviation > 0.5% (review required)."
        ))
        for col_idx, header in enumerate(CHECK_COLS, start=1):
            ws_chk.cell(row=2, column=col_idx, value=header)
        for row_idx, row_data in enumerate(
                check_out.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws_chk.cell(row=row_idx, column=col_idx, value=value)

        # Sheet 3: Empty placeholder
        writer.book.create_sheet('New Portfolio Allocation >>')

        # Sheets 4+: Relevant allocation tabs
        for sheet in relevant_sheets:
            if sheet in new_xls.sheet_names:
                prepare_smart_tab(new_xls, sheet).to_excel(
                    writer, sheet_name=sheet, index=False)

    # Re-open for formatting
    wb = load_workbook(path)

    fmt_sheet(wb['Trades'], HEADER_FILL_TRADE)

    ws_c = wb['Check']
    note_cell = ws_c.cell(row=1, column=1)
    note_cell.font      = Font(bold=True, italic=True, name="Arial",
                                size=10, color="375623")
    note_cell.alignment = Alignment(horizontal='left')
    ws_c.merge_cells(start_row=1, start_column=1,
                     end_row=1, end_column=len(CHECK_COLS))
    fmt_sheet(ws_c, HEADER_FILL_CHECK, freeze='A3',
              header_row=2, data_start_row=3)

    v_idx = CHECK_COLS.index('Switch Out Fund Weight (%) after Rebalance') + 1
    w_idx = CHECK_COLS.index('Target Allocation of Switch Out Fund (%)') + 1
    apply_vw_conditional(ws_c, v_idx, w_idx, start_row=3)

    # Format U as %, V as %
    u_idx    = CHECK_COLS.index('Total Switch Out %') + 1
    u_letter = get_column_letter(u_idx)
    v_letter = get_column_letter(v_idx)
    for i in range(3, len(trades_out) + 3):
        ws_c[f'{u_letter}{i}'].number_format = '0.00%'
        ws_c[f'{v_letter}{i}'].number_format = '0.00%'

    for sheet in relevant_sheets:
        if sheet in wb.sheetnames:
            fmt_sheet(wb[sheet], HEADER_FILL_ALLOC)

    # Sheet order
    desired = (['Trades', 'Check', 'New Portfolio Allocation >>'] +
               relevant_sheets)
    wb._sheets.sort(key=lambda s: desired.index(s.title)
                    if s.title in desired else 999)

    wb.save(path)
    print(f"  Saved: {path}")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("Switch Order Generator — initialising...")
    time.sleep(0.5)

    show_msg("Step 1", "Select the REBALANCE INPUT file.")
    rebalance_path = prompt_file("Step 1: Rebalance Input File")
    if not rebalance_path:
        sys.exit()

    show_msg("Step 2", "Select the NEW ALLOCATION file.")
    new_alloc_path = prompt_file("Step 2: New Allocation File")
    if not new_alloc_path:
        sys.exit()

    show_msg("Step 3", "Select the OLD ALLOCATION file.")
    old_alloc_path = prompt_file("Step 3: Old Allocation File")
    if not old_alloc_path:
        sys.exit()

    print("Loading files...")
    rebal_df = clean_df(load_excel_safe(rebalance_path))
    new_xls  = pd.ExcelFile(new_alloc_path)
    old_xls  = pd.ExcelFile(old_alloc_path)

    rebal_df['svc_acct']  = rebal_df['svc_acct'].astype(str).str.strip()
    rebal_df['port_type'] = rebal_df['port_type'].astype(str).str.strip().str.replace(
        r'^T_', '', regex=True)

    # Identify portfolio types that appear in both rebal input and both alloc files
    rebal_port_types  = set(rebal_df['port_type'].unique())
    common_port_types = rebal_port_types & set(new_xls.sheet_names) & set(old_xls.sheet_names)

    # Filter to only port types with actual changes
    changed_port_types = []
    for pt in sorted(common_port_types):
        deltas = compute_deltas(new_xls, old_xls, pt)
        if deltas is not None and not deltas.empty:
            changed_port_types.append(pt)

    if not changed_port_types:
        show_msg("No Changes", "No allocation changes detected between old and new files.")
        sys.exit()

    print(f"  Portfolio types with changes: {changed_port_types}")

    # ── Step 4: Generate mapping suggestion and ask user to review ──
    show_msg("Step 4 — Mapping Review",
             "The script will now generate a suggested fund switch mapping file.\n\n"
             "Please:\n"
             "1. Save it to a convenient location\n"
             "2. Open it and verify/edit the pairings\n"
             "3. Save the file\n"
             "4. Click OK in the next prompt to continue")

    show_msg("Step 4a", "Select a folder to save the SUGGESTED MAPPING file.")
    mapping_folder = prompt_folder("Step 4a: Select folder for mapping file")
    if not mapping_folder:
        sys.exit()

    today_str    = dt.datetime.today().strftime("%Y%m%d")
    mapping_path = os.path.join(mapping_folder,
                                f"SwitchMapping_SUGGESTED_{today_str}.xlsx")

    print("Generating suggested mapping...")
    write_mapping_suggestion(mapping_path, new_xls, old_xls, changed_port_types)

    show_msg("Action Required",
             f"Suggested mapping saved to:\n{mapping_path}\n\n"
             "Please open, review and edit the mapping file now.\n"
             "Make sure to SAVE it before clicking OK.")

    # ── Step 5: Load confirmed mapping ──
    show_msg("Step 5", "Select the CONFIRMED MAPPING file (the one you just reviewed).")
    confirmed_mapping_path = prompt_file(
        "Step 5: Select Confirmed Mapping File",
        filetypes=[("Excel files", "*.xlsx *.xlsm")])
    if not confirmed_mapping_path:
        sys.exit()

    mapping_df = load_mapping(confirmed_mapping_path)
    if mapping_df.empty:
        show_msg("Error", "No valid mapping rows found. Please check the mapping file.")
        sys.exit()

    print(f"  Loaded {len(mapping_df)} mapping rows across "
          f"{mapping_df['port_type'].nunique()} portfolio types.")

    # ── Build fund name map ──
    print("Building fund name map...")
    fund_name_map = build_fund_name_map([new_xls, old_xls])

    # ── Compute trades ──
    print("Computing switch trades...")
    trades_df = compute_switch_trades(
        rebal_df, new_xls, old_xls, mapping_df, fund_name_map)

    if trades_df.empty:
        show_msg("No Trades",
                 "No trades generated. Check that mapping file port_types match "
                 "the rebalance input and allocation files.")
        sys.exit()

    print(f"  {len(trades_df)} trade rows generated.")
    pairs = trades_df.groupby(
        ['Portfolio Type', 'MgrCd', 'FundCd', 'SIMgrCd', 'SIFundCd']).size()
    print("\nSwitch pairs generated:")
    print(pairs.to_string())

    # Relevant allocation sheets = port types in trades
    relevant_sheets = [s for s in new_xls.sheet_names
                       if s in trades_df['Portfolio Type'].unique()]

    # ── Step 6: Output folder ──
    show_msg("Step 6", "Select a folder to save the SWITCH ORDER output file.")
    output_folder = prompt_folder("Step 6: Output Folder")
    if not output_folder:
        sys.exit()

    output_path = os.path.join(output_folder, f"SwitchingOrder_{today_str}.xlsx")

    print("\nWriting output file...")
    write_switch_file(
        path            = output_path,
        trades_df       = trades_df,
        new_xls         = new_xls,
        relevant_sheets = relevant_sheets,
    )

    print(f"\nDone. Output saved to:\n  {output_path}")
    sys.exit()
