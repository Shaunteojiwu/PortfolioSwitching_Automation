import os
import sys
import time
import pandas as pd
import datetime as dt
import warnings
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import io

warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────
# FORMATTING CONSTANTS
# ─────────────────────────────────────────────
HEADER_FILL_TRADE   = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_CHECK   = PatternFill("solid", fgColor="375623")
HEADER_FILL_ALLOC   = PatternFill("solid", fgColor="7B2C2C")
HEADER_FILL_MAPPING = PatternFill("solid", fgColor="7B5B00")
HEADER_FONT         = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT           = Font(name="Arial", size=10)
ALT_FILL_LIGHT      = PatternFill("solid", fgColor="EBF3FB")
ALT_FILL_NONE       = PatternFill("solid", fgColor="FFFFFF")
OK_FILL             = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL           = PatternFill("solid", fgColor="FFF2CC")
SUGGEST_FILL        = PatternFill("solid", fgColor="FFF9C4")
THIN                = Side(style='thin', color="D9D9D9")
BORDER              = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)




# ─────────────────────────────────────────────
# HELPER FUNCTION
# ─────────────────────────────────────────────

def create_root():
    root = tk.Tk() # creates a temporary tkinter window, which is needed for the file dialog to work
    root.withdraw()  # Hides the empty root window from the user, but it still exists in the background.
    root.lift() # Brings the hidden window (and its dialogs) to the front.
    root.attributes("-topmost", True) # Ensures dialogs appear above other windows.
    return root

# ─────────────────────────────────────────────
# UI HELPERS
# ─────────────────────────────────────────────
def show_msg(title, message):
    root =create_root() #creates temp tinker window, hides it, force popup to appear infront of other windows
    messagebox.showinfo(title, message, parent=root); root.destroy() #shows an information message, close the temporary window

def prompt_file(title, filetypes=None):
    root =create_root()
    path = filedialog.askopenfilename(title=title, filetypes=filetypes or []) #user selects excel file
    root.destroy(); return path #close the temporary window and return file path

def prompt_folder(title):
    root =create_root()
    path = filedialog.askdirectory(title=title); root.destroy(); return path #opens a folder, close window and return path

def ask_yes_no(title, message):
    root =create_root()
    result = messagebox.askyesno(title, message, parent=root); root.destroy() # show a window where yes==True, no=False, close windowand retunr result
    return result

# ─────────────────────────────────────────────
# FILE LOADING
# ─────────────────────────────────────────────
def convert_html_xls_to_xlsx(input_path):
    with open(input_path, 'rb') as f: #open file and read as binary
        content = f.read() #reads all the html content
    soup = BeautifulSoup(content, 'html.parser') #reads the html and creates a tree (HTML Text -> Beautiful Soup object),can find tables, rows, cells, etc.
    tables = pd.read_html(io.StringIO(str(soup)), header=None) #turns object back into html,create a fake memory in file, pandas then read the html and search for <table> and converts it to df, [df1,df2,df3]
    if not tables:
        raise ValueError("No HTML tables found.") #if no tables found
    raw_df = tables[0] # read first table
    raw_df.columns = raw_df.iloc[0].astype(str).str.strip() #rename column as first row, this cause repeated column names (coumn index and row index 0)
    df = raw_df.iloc[1:].reset_index(drop=True) # take from 1st row onwards, to remove row 0 that contain header, reset the index to start from index 0 and drop the index column
    output_path = os.path.splitext(input_path)[0] + ".xlsx" # split the input file into 2 parts (remove xls) and add xlsx to the 2nd part
    df.to_excel(output_path, index=False) # convert df to excel and output to the output path
    return output_path

def load_excel_safe(path):
    ext = os.path.splitext(path)[-1].lower()
    try:
        if ext == '.xls':
            try:
                return pd.read_excel(path, sheet_name=0, engine='xlrd')
            except Exception:
                return pd.read_excel(convert_html_xls_to_xlsx(path),
                                     sheet_name=0, engine='openpyxl')
        elif ext in ['.xlsx', '.xlsm']:
            return pd.read_excel(path, sheet_name=0, engine='openpyxl')
        else:
            raise ValueError(f"Unsupported extension: {ext}")
    except Exception as e:
        print(f"Failed to read: {path}\n{e}"); sys.exit()

def clean_df(df):
    df.columns = [str(c).strip() for c in df.columns]
    return df

# ─────────────────────────────────────────────
# ALLOCATION HELPERS
# ─────────────────────────────────────────────
def load_allocation(xls, sheet): #xls is excel file object, old and new excel file in our case created in main
    """
    Load allocation sheet. Handles:
    - 'weights' or 'weights (%)' column name
    - Decimal (0.20) or % integer (20) format — normalises to decimal
    - Zero-pads mgr_code and fund_code to 3 digits
    """
    try:
        df = pd.read_excel(xls, sheet_name=sheet).rename(
            columns=lambda x: str(x).strip().lower())
    except Exception:
        return None
    if 'mgr_code' not in df.columns or 'fund_code' not in df.columns:
        return None
    # Accept 'weights (%)' as 'weights'
    if 'weights' not in df.columns and 'weights (%)' in df.columns:
        df.rename(columns={'weights (%)': 'weights'}, inplace=True)
    if 'weights' not in df.columns:
        return None
    df['mgr_code'] = df['mgr_code'].apply(
        lambda x: str(int(float(x))).zfill(3) #take the mgr code and always make it 3 values, eg, "001","012","123"
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '') # if not empty run that fnc otherwise return ""
    df['fund_code'] = df['fund_code'].apply(
        lambda x: str(int(float(x))).zfill(3)
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
    df['weights'] = pd.to_numeric(df['weights'], errors='coerce').fillna(0) # change all weights to numeric, fill up na with 0
    # Normalise % integers to decimals
    if df['weights'].max() > 1: # this weight percentage, must be in decimal
        df['weights'] = df['weights'] / 100 
    df['key'] = df['mgr_code'] + df['fund_code']
    return df

def prepare_smart_tab(xls, sheet):
    """Display-ready allocation df: weights as %, fund_code zero-padded."""
    df = pd.read_excel(xls, sheet_name=sheet, dtype={'fund_code': str})
    df.columns = [c.strip() for c in df.columns]
    if 'fund_code' in df.columns:
        df['fund_code'] = df['fund_code'].apply(
            lambda x: str(int(float(x))).zfill(3) if pd.notna(x) else x)
    weight_col = 'weights (%)' if 'weights (%)' in df.columns else 'weights'
    if weight_col in df.columns:
        df[weight_col] = pd.to_numeric(df[weight_col], errors='coerce')
        if df[weight_col].dropna().le(1).all(): #drop missing values, if weight less than equal to 1, check if all is True (less/equal to 1)
            df[weight_col] = (df[weight_col] * 100).round(2) # make it a percentage
        if weight_col == 'weights':
            df.rename(columns={'weights': 'weights (%)'}, inplace=True) # reanme it to percentage, for output (human readability)
    return df

# ─────────────────────────────────────────────
# FUND NAME MAP
# ─────────────────────────────────────────────
def build_fund_name_map(xls_list): #old and new excel file
    fmap = {}
    for xls in xls_list:
        for sheet in xls.sheet_names: #read through all the sheet within
            try:
                df = pd.read_excel(xls, sheet_name=sheet).rename(
                    columns=lambda x: str(x).strip().lower()) #read the file anf sheet, and lowercase,strip the column heaers
                if 'mgr_code' not in df.columns or 'fund_code' not in df.columns:
                    continue
                for _, r in df.iterrows(): #iterrows, rows index and the row
                    try:
                        mgr  = str(int(float(r['mgr_code']))).zfill(3)
                        fund = str(int(float(r['fund_code']))).zfill(3)
                        name = str(r.get('fund name', '')) \
                               if pd.notna(r.get('fund name', '')) else ''
                        if name:
                            fmap[(mgr, fund)] = name # {("001", "005"): "PIMCO Income Fund"}
                    except (ValueError, TypeError):
                        continue
            except Exception:
                continue
    return fmap

# ─────────────────────────────────────────────
# DELTA CALCULATION
# ─────────────────────────────────────────────
def compute_deltas(new_xls, old_xls, port_type):
    """
    Returns merged df with columns:
      mgr_code, fund_code, key, old_w, new_w, delta, fund_name
    Only returns rows where delta != 0.
    """
    new_alloc = load_allocation(new_xls, port_type)
    old_alloc = load_allocation(old_xls, port_type)
    if new_alloc is None or old_alloc is None:
        return None

    new_cols = new_alloc[['key', 'mgr_code', 'fund_code', 'weights']].rename(
        columns={'weights': 'new_w', 'mgr_code': 'mgr_new', 'fund_code': 'fund_new'})
    old_cols = old_alloc[['key', 'mgr_code', 'fund_code', 'weights']].rename(
        columns={'weights': 'old_w', 'mgr_code': 'mgr_old', 'fund_code': 'fund_old'})

    merged = pd.merge(new_cols, old_cols, on='key', how='outer') #match by key and keep everything
    merged['new_w']     = pd.to_numeric(merged['new_w'], errors='coerce').fillna(0)
    merged['old_w']     = pd.to_numeric(merged['old_w'], errors='coerce').fillna(0)
    merged['mgr_code']  = merged['mgr_new'].fillna(merged['mgr_old'])
    merged['fund_code'] = merged['fund_new'].fillna(merged['fund_old'])
    merged['delta']     = (merged['new_w'] - merged['old_w']).round(6)

    # Attach fund names from both files
    name_map = build_fund_name_map([new_xls, old_xls])
    merged['fund_name'] = merged.apply(
        lambda r: name_map.get((r['mgr_code'], r['fund_code']), ''), axis=1) #attach fund names  "PIMCO Income Fund"

    return merged[merged['delta'] != 0].copy() # return all thsose have changes

# ─────────────────────────────────────────────
# CORE TRADE CALCULATION
# ─────────────────────────────────────────────
def compute_switch_trades(rebal_df, new_xls, old_xls, fund_name_map):
    """
    Pure many-to-many: every switch-out fund splits proportionally across
    ALL switch-in funds by their relative positive deltas.
      units_to_fund_B = units_out x (delta_B / total_positive_delta)
    Last destination gets remainder to eliminate rounding residuals.
    """
    rebal_df = rebal_df.copy() #rebalance sheet copy
    rebal_df.columns = [c.strip() for c in rebal_df.columns]
    rebal_df['svc_acct']  = rebal_df['svc_acct'].astype(str).str.strip().str.zfill(7)
    rebal_df['port_type'] = rebal_df['port_type'].astype(str).str.strip().str.replace(
        r'^T_', '', regex=True) #^ start of string T_ replace it with empty string
    rebal_df['code'] = rebal_df['code'].apply(
        lambda x: str(int(float(x))).zfill(6)
        if pd.notna(x) and str(x).strip() not in ('', 'nan') else '')
    rebal_df['MgrCd']    = rebal_df['code'].str[:3] # take the first 3 elements of string
    rebal_df['FundCd']   = rebal_df['code'].str[-3:] # take last 3 elements of string
    rebal_df['qty']      = pd.to_numeric(rebal_df['qty'],      errors='coerce').fillna(0)
    rebal_df['mkt_val']  = pd.to_numeric(rebal_df['mkt_val'],  errors='coerce').fillna(0)
    rebal_df['total_pf'] = pd.to_numeric(rebal_df['total_pf'], errors='coerce').fillna(0)

    # Build rebal lookup: (acct, mgr, fund) -> row
    rebal_lookup = {}
    for _, r in rebal_df.iterrows():
        rebal_lookup[(r['svc_acct'], r['MgrCd'], r['FundCd'])] = r



    today_str  = dt.datetime.today().strftime("%Y%m%d")
    trade_rows = []

    valid_port_types = set(new_xls.sheet_names) & set(old_xls.sheet_names)

    for port_type in rebal_df['port_type'].unique():
        if port_type not in valid_port_types:
            continue

        # Compute deltas for this portfolio type
        deltas = compute_deltas(new_xls, old_xls, port_type)
        if deltas is None or deltas.empty:
            continue

        switch_out = deltas[deltas['delta'] < 0].copy()
        switch_in  = deltas[deltas['delta'] > 0].copy()
        if switch_out.empty or switch_in.empty:
            continue

        total_positive_delta = switch_in['delta'].sum()
        total_negative_delta = abs(switch_out['delta'].sum())
        
          # Optional safety check
        if round(total_positive_delta, 6) != round(total_negative_delta, 6):
            print(
                f"WARNING [{port_type}]: positive delta "
                f"{total_positive_delta:.6f} does not equal negative delta "
                f"{total_negative_delta:.6f}"
            )
            
        in_list = switch_in.to_dict('records')  # "records' in built method df to list of dicts
        n_in    = len(in_list)

        fdsrc  = 'SRS-IA' if port_type.endswith('_SRS') else 'Cash'
        divopt = 'Withdraw' if port_type.startswith('MD') else 'Reinvest'

        old_alloc = load_allocation(old_xls, port_type)
        new_alloc = load_allocation(new_xls, port_type)
        old_w_map = {r['mgr_code'] + r['fund_code']: r['weights']
                     for _, r in old_alloc.iterrows()}
        new_w_map = {r['mgr_code'] + r['fund_code']: r['weights']
                     for _, r in new_alloc.iterrows()}

        unique_accts = rebal_df[
            rebal_df['port_type'] == port_type]['svc_acct'].unique()

        for acct_no in unique_accts:
            for _, out_row in switch_out.iterrows():
                mgr_out  = out_row['mgr_code']
                fund_out = out_row['fund_code']
                key      = mgr_out + fund_out

                rebal_row = rebal_lookup.get((acct_no, mgr_out, fund_out))
                if rebal_row is None:
                    continue
                qty      = rebal_row['qty']
                mkt_val  = rebal_row['mkt_val']
                total_pf = rebal_row['total_pf']
                if qty <= 0:
                    continue

                old_w = old_w_map.get(key, 0)
                new_w = new_w_map.get(key, 0)

                # Units switched out
                if new_w == 0:
                    units_out = qty                     # full switch
                elif old_w > 0:
                    units_out = round(((old_w - new_w) / old_w) * qty, 3) # partial reduction
                else:
                    continue
                if units_out <= 0:
                    continue

                # SO check columns — same for every SI destination row
                r_val   = round(qty, 3)               # SO: Units before Rebal
                s_val   = round(units_out, 3)          # Total Units SO
                t_val   = round(qty - units_out, 3)    # SO: Units Remaining
                u_val   = round(mkt_val, 2)            # SO: MV before Rebal (SGD)
                v_val   = round(total_pf, 2)           # Account Portfolio AUM
                # W: SO Fund Weight after Rebal = ((T/R)*U)/V * 100
                w_val   = round(((t_val / r_val) * u_val) / v_val * 100, 2) \
                          if (r_val > 0 and v_val > 0) else 0.0
                x_val   = round(new_w * 100, 2)        # Target alloc SO (%)
                y_val   = round(w_val - x_val, 2)      # Diff W-X
                so_name = fund_name_map.get((mgr_out, fund_out), '')
                
                # Split proportionally across ALL switch-in funds
                allocated = 0.0
                for i, in_row in enumerate(in_list):
                    is_last    = (i == n_in - 1)
                    mgr_in     = in_row['mgr_code']
                    fund_in    = in_row['fund_code']
                    delta_in   = in_row['delta']
                    si_name    = fund_name_map.get((mgr_in, fund_in), '')

                    if is_last:
                        units_this = round(units_out - allocated, 3)
                    else:
                        units_this  = round(
                            units_out * (delta_in / total_positive_delta), 3)
                        allocated  += units_this
                        
                    if units_this <= 0:
                        continue

                    order_type = 'SO' if mgr_out == mgr_in else 'ESO'

                    trade_rows.append({
                        'AccountNo':   acct_no,
                        'OrderType':   order_type,
                        'MgrCd':       mgr_out,
                        'FundCd':      fund_out,
                        'FdSrc':       fdsrc,
                        'SIMgrCd':     mgr_in,
                        'SIFundCd':    fund_in,
                        'NetSalesChg': 0,
                        'SwChg':       0,
                        'DivOpt':      divopt,
                        'Units':       units_this,
                        'FACode':      '',
                        'TransDate':   today_str,
                        'Portfolio Type':                               port_type,
                        'Switch Out Fund Name':                         so_name,
                        'Switch In Fund Name':                          si_name,
                        'SO: Units before Rebalance':                   r_val,
                        'Total Units SO':                               s_val,
                        'SO: Units Remaining after Rebalance':          t_val,
                        'SO: MV of Fund before Rebalance (SGD)':        u_val,
                        'Account Portfolio Amount (SGD)':               v_val,
                        'SO Fund Weight (%) after Rebalance':           w_val,
                        'Target Allocation of Switch Out Fund (%)':     x_val,
                        'Diff':                                         y_val,
                    })


    return pd.DataFrame(trade_rows) if trade_rows else pd.DataFrame()


# ─────────────────────────────────────────────
# EXCEL FORMATTING
# ─────────────────────────────────────────────
def fmt_sheet(ws, header_fill, freeze='A2', header_row=1, data_start_row=2):
    for cell in ws[header_row]:
        cell.fill      = header_fill
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border    = BORDER
    ws.row_dimensions[header_row].height = 30
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row),
                                  start=data_start_row):
        fill = ALT_FILL_LIGHT if row_idx % 2 == 0 else ALT_FILL_NONE
        for cell in row:
            cell.font   = BODY_FONT
            cell.border = BORDER
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = fill
    if freeze:
        ws.freeze_panes = freeze
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0)
                      for c in col)
        # col[0] may be a MergedCell — find first real cell with column_letter
        first = next((c for c in col if hasattr(c, 'column_letter')), None)
        if first:
            ws.column_dimensions[first.column_letter].width = min(
                max(max_len + 2, 10), 50)

def apply_vw_conditional(ws, v_col_idx, w_col_idx, start_row=2):
    for row in ws.iter_rows(min_row=start_row):
        v_cell = row[v_col_idx - 1]
        w_cell = row[w_col_idx - 1]
        if v_cell.value is not None and w_cell.value is not None:
            try:
                v_pct = float(v_cell.value) * 100
                w_pct = float(w_cell.value)
                v_cell.fill = OK_FILL if abs(v_pct - w_pct) < 0.5 else WARN_FILL
            except (TypeError, ValueError):
                pass

# ─────────────────────────────────────────────
# OUTPUT FILE WRITER
# ─────────────────────────────────────────────
def write_switch_file(path, trades_df, new_xls, relevant_sheets):
    TRADE_COLS = ['AccountNo', 'OrderType', 'MgrCd', 'FundCd',
                  'FdSrc', 'SIMgrCd', 'SIFundCd', 'NetSalesChg', 'SwChg',
                  'DivOpt', 'Units', 'FACode', 'TransDate']
    CHECK_COLS = TRADE_COLS + [
        'Portfolio Type',                               # N
        'Switch Out Fund Name',                         # O
        'Switch In Fund Name',                          # P
        'SO: Units before Rebalance',                   # Q
        'Total Units SO',                               # R
        'SO: Units Remaining after Rebalance',          # S
        'SO: MV of Fund before Rebalance (SGD)',        # T
        'Account Portfolio Amount (SGD)',               # U
        'SO Fund Weight (%) after Rebalance',           # V
        'Target Allocation of Switch Out Fund (%)',     # W
        'Diff',                                         # X
    ]

    trades_out = trades_df[TRADE_COLS].copy()
    check_out  = trades_df[CHECK_COLS].copy()

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        # Sheet 1: Trades — clean values
        trades_out.to_excel(writer, sheet_name='Trades', index=False)

        # Sheet 2: Check — note row 1, headers row 2, data from row 3
        ws_chk = writer.book.create_sheet('Check')
        writer.sheets['Check'] = ws_chk
        ws_chk.cell(row=1, column=1, value=(
            "CHECK TAB — Green = Diff (col X) within ±1% of target.  "
            "Yellow = deviation > 1% (review required).  V = SO weight after rebal; W = SO target."
            "W/X = Switch Out fund weight vs target; Z/AA = Switch In fund weight vs target."
        ))
        for col_idx, header in enumerate(CHECK_COLS, start=1):
            ws_chk.cell(row=2, column=col_idx, value=header)
        for row_idx, row_data in enumerate(
                check_out.itertuples(index=False), start=3):
            for col_idx, value in enumerate(row_data, start=1):
                ws_chk.cell(row=row_idx, column=col_idx, value=value)

        # Sheet 3: Empty placeholder
        writer.book.create_sheet('New Portfolio Allocation >>')

        # Sheets 4+: Relevant allocation tabs
        for sheet in relevant_sheets:
            if sheet in new_xls.sheet_names:
                prepare_smart_tab(new_xls, sheet).to_excel(
                    writer, sheet_name=sheet, index=False)

    # Re-open for formatting
    wb = load_workbook(path)

    fmt_sheet(wb['Trades'], HEADER_FILL_TRADE)

    ws_c = wb['Check']
    note_cell = ws_c.cell(row=1, column=1)
    note_cell.font      = Font(bold=True, italic=True, name="Arial",
                                size=10, color="375623")
    note_cell.alignment = Alignment(horizontal='left')
    ws_c.merge_cells(start_row=1, start_column=1,
                     end_row=1, end_column=len(CHECK_COLS))
    fmt_sheet(ws_c, HEADER_FILL_CHECK, freeze='A3',
              header_row=2, data_start_row=3)

    # Bold row 2 header with dark green fill
    for cell in ws_c[2]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill = HEADER_FILL_CHECK
        cell.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)
        cell.border = BORDER
    ws_c.row_dimensions[2].height = 30

    # Conditional formatting on X (Diff) — ±1% threshold
    x_col  = get_column_letter(CHECK_COLS.index('Diff') + 1)
    n_data = len(trades_out)
    for i in range(3, n_data + 3):
        cell = ws_c[f'{x_col}{i}']
        try:
            val = float(cell.value) if cell.value is not None else None
            if val is not None:
                cell.fill = OK_FILL if abs(val) <= 1.0 else WARN_FILL
        except (TypeError, ValueError):
            pass


    for sheet in relevant_sheets:
        if sheet in wb.sheetnames:
            fmt_sheet(wb[sheet], HEADER_FILL_ALLOC)

    # Sheet order
    desired = (['Trades', 'Check', 'New Portfolio Allocation >>'] +
               relevant_sheets)
    wb._sheets.sort(key=lambda s: desired.index(s.title)
                    if s.title in desired else 999)

    wb.save(path)
    print(f"  Saved: {path}")

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    print("Switch Order Generator — initialising...")
    time.sleep(0.5)

    show_msg("Step 1", "Select the REBALANCE INPUT file.")
    rebalance_path = prompt_file("Step 1: Rebalance Input File")
    if not rebalance_path:
        sys.exit()

    show_msg("Step 2", "Select the NEW ALLOCATION file.")
    new_alloc_path = prompt_file("Step 2: New Allocation File")
    if not new_alloc_path:
        sys.exit()

    show_msg("Step 3", "Select the OLD ALLOCATION file.")
    old_alloc_path = prompt_file("Step 3: Old Allocation File")
    if not old_alloc_path:
        sys.exit()

    print("Loading files...")
    rebal_df = clean_df(load_excel_safe(rebalance_path))
    new_xls  = pd.ExcelFile(new_alloc_path) #creates an excel file object, which can be used to read multiple sheets
    old_xls  = pd.ExcelFile(old_alloc_path)

    rebal_df['svc_acct']  = rebal_df['svc_acct'].astype(str).str.strip()
    rebal_df['port_type'] = rebal_df['port_type'].astype(str).str.strip().str.replace(
        r'^T_', '', regex=True)

    # Identify portfolio types that appear in both rebal input and both alloc files
    rebal_port_types  = set(rebal_df['port_type'].unique())
    common_port_types = rebal_port_types & set(new_xls.sheet_names) & set(old_xls.sheet_names) #return those that intersect in all 3 sets

    # Filter to only port types with actual changes
    changed_port_types = []
    for pt in sorted(common_port_types): #alphabetical order
        deltas = compute_deltas(new_xls, old_xls, pt)
        if deltas is not None and not deltas.empty:
            changed_port_types.append(pt)

    if not changed_port_types:
        show_msg("No Changes", "No allocation changes detected between old and new files.")
        sys.exit()

    print(f"  Portfolio types with changes: {changed_port_types}")

    # ── Build fund name map ──
    print("Building fund name map...")
    fund_name_map = build_fund_name_map([new_xls, old_xls])

    today_str = dt.datetime.today().strftime("%Y%m%d")

    # ── Compute trades ──
    print("Computing switch trades...")
    trades_df = compute_switch_trades(rebal_df, new_xls, old_xls, fund_name_map)

    if trades_df.empty:
        show_msg("No Trades",
                 "No trades generated. Check that the old and new allocation "
                 "files have matching sheet names and different weights.")
        sys.exit()

    print(f"  {len(trades_df)} trade rows generated.")
    pairs = trades_df.groupby(
        ['Portfolio Type', 'MgrCd', 'FundCd', 'SIMgrCd', 'SIFundCd']).size()
    print("\nSwitch pairs generated:")
    print(pairs.to_string())

    # Relevant allocation sheets = port types in trades
    relevant_sheets = [s for s in new_xls.sheet_names
                       if s in trades_df['Portfolio Type'].unique()]

    # ── Step 6: Output folder ──
    show_msg("Step 4", "Select a folder to save the SWITCH ORDER output file.")
    output_folder = prompt_folder("Step 4: Output Folder")
    if not output_folder:
        sys.exit()

    output_path = os.path.join(output_folder, f"SwitchingOrder_{today_str}.xlsx")

    print("\nWriting output file...")
    write_switch_file(
        path            = output_path,
        trades_df       = trades_df,
        new_xls         = new_xls,
        relevant_sheets = relevant_sheets,
    )

    print(f"\nDone. Output saved to:\n  {output_path}")
    sys.exit()
