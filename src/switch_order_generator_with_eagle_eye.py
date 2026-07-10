import os
import sys
import time
import io
import warnings
import datetime as dt

import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ------------------------------------------------------------
# FORMATTING CONSTANTS
# ------------------------------------------------------------
HEADER_FILL_TRADE = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_CHECK = PatternFill("solid", fgColor="375623")
HEADER_FILL_ALLOC = PatternFill("solid", fgColor="7B2C2C")
HEADER_FILL_EAGLE = PatternFill("solid", fgColor="7030A0")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BODY_FONT = Font(name="Arial", size=10)
ALT_FILL_LIGHT = PatternFill("solid", fgColor="EBF3FB")
ALT_FILL_NONE = PatternFill("solid", fgColor="FFFFFF")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# ------------------------------------------------------------
# UI HELPERS
# ------------------------------------------------------------
def create_root():
    root = tk.Tk()
    root.withdraw()
    root.lift()
    root.attributes("-topmost", True)
    return root


def show_msg(title, message):
    root = create_root()
    try:
        messagebox.showinfo(title, message, parent=root)
    finally:
        root.destroy()


def prompt_file(title, filetypes=None):
    root = create_root()
    try:
        return filedialog.askopenfilename(
            title=title,
            filetypes=filetypes or [],
            parent=root,
        )
    finally:
        root.destroy()


def prompt_folder(title):
    root = create_root()
    try:
        return filedialog.askdirectory(title=title, parent=root)
    finally:
        root.destroy()


def ask_yes_no(title, message):
    root = create_root()
    try:
        return messagebox.askyesno(title, message, parent=root)
    finally:
        root.destroy()


# ------------------------------------------------------------
# FILE LOADING
# ------------------------------------------------------------
def convert_html_xls_to_xlsx(input_path):
    with open(input_path, "rb") as f:
        content = f.read()

    soup = BeautifulSoup(content, "html.parser")
    tables = pd.read_html(io.StringIO(str(soup)), header=None)
    if not tables:
        raise ValueError("No HTML tables found.")

    raw_df = tables[0]
    raw_df.columns = raw_df.iloc[0].astype(str).str.strip()
    df = raw_df.iloc[1:].reset_index(drop=True)

    output_path = os.path.splitext(input_path)[0] + ".xlsx"
    df.to_excel(output_path, index=False)
    return output_path


def load_excel_safe(path):
    ext = os.path.splitext(path)[-1].lower()
    try:
        if ext == ".xls":
            try:
                return pd.read_excel(path, sheet_name=0, engine="xlrd")
            except Exception:
                converted = convert_html_xls_to_xlsx(path)
                return pd.read_excel(converted, sheet_name=0, engine="openpyxl")
        if ext in (".xlsx", ".xlsm"):
            return pd.read_excel(path, sheet_name=0, engine="openpyxl")
        raise ValueError(f"Unsupported extension: {ext}")
    except Exception as exc:
        print(f"Failed to read: {path}\n{exc}")
        sys.exit(1)


def clean_df(df):
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def normalise_account_number(series):
    """Return account numbers as clean 7-character strings."""
    return (
        series.astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.zfill(7)
    )


# ------------------------------------------------------------
# STANDARD PORTFOLIO-TYPE FUND SOURCE MAPPING
# ------------------------------------------------------------
PORT_TYPE_FUND_SOURCE = {
    "HVP1": "Cash",
    "HVP2": "Cash",
    "HVP3": "Cash",
    "HVP_Income": "Cash",
    "HVP_Thematic": "Cash",
    "HVP Thematic": "Cash",
    "CBP": "Cash",
    "HVP2_SRS": "SRS-IA",
    "HVP3_SRS": "SRS-IA",
    "CBP_SRS": "SRS-IA",
}


# ------------------------------------------------------------
# EAGLE EYE HELPERS - NEW
# ------------------------------------------------------------
def prepare_eagle_eye_export(eagle_df):
    """
    Clean the Eagle Eye account export and retain one row per service account.

    Required columns:
      - Svc Acct No
      - Fund Source

    Other Eagle Eye columns are retained and can be copied into the separate
    Eagle Eye output for checking/reference.
    """
    eagle = clean_df(eagle_df)

    required = {"Svc Acct No", "Fund Source"}
    missing = required - set(eagle.columns)
    if missing:
        raise ValueError(
            "Eagle Eye file is missing required column(s): "
            + ", ".join(sorted(missing))
        )

    eagle["Svc Acct No"] = normalise_account_number(eagle["Svc Acct No"])
    eagle["Fund Source"] = eagle["Fund Source"].fillna("").astype(str).str.strip()

    # Validate that a service account does not map to conflicting fund sources.
    conflicts = (
        eagle.loc[eagle["Fund Source"].ne("")]
        .groupby("Svc Acct No")["Fund Source"]
        .nunique()
    )
    conflict_accounts = conflicts[conflicts > 1].index.tolist()
    if conflict_accounts:
        preview = ", ".join(conflict_accounts[:10])
        raise ValueError(
            "Conflicting Fund Source values found for Eagle Eye account(s): "
            f"{preview}"
        )

    # Keep the first record per account after validation.
    return eagle.drop_duplicates(subset=["Svc Acct No"], keep="first").copy()


def build_eagle_eye_lookup(eagle_df):
    """
    Convert the cleaned Eagle Eye table into:

        {
            '1234567': {
                'Fund Source': 'Cash',
                'FA': '...',
                'Client No': '...',
                ...
            }
        }
    """
    eagle = prepare_eagle_eye_export(eagle_df)
    return eagle.set_index("Svc Acct No").to_dict(orient="index")


def create_eagle_eye_output(trades_df, eagle_lookup):
    """
    Create a separate Eagle Eye-enriched output.

    The switching calculations are unchanged. This function only joins
    account-level Eagle Eye information onto the already-generated trades.
    """
    output = trades_df.copy()
    output["AccountNo"] = normalise_account_number(output["AccountNo"])

    preferred_eagle_columns = [
        "Start Date",
        "Client No",
        "Name",
        "MM",
        "FA",
        "Fund Source",
        "HPR%",
        "Capital Holding(SGD)",
        "Market Value(SGD)",
        "Ann TWR%",
        "Mthly TWR%",
        "Yearly TWR%",
        "Capital Changed(SGD)",
        "% Inv",
    ]

    # Add only columns that actually exist in the selected Eagle Eye export.
    available_columns = set()
    for account_data in eagle_lookup.values():
        available_columns.update(account_data.keys())

    columns_to_add = [
        column for column in preferred_eagle_columns if column in available_columns
    ]

    for column in columns_to_add:
        output[f"Eagle Eye - {column}"] = output["AccountNo"].map(
            lambda acct: eagle_lookup.get(acct, {}).get(column, "")
        )

    # Override the operational FdSrc only in the Eagle Eye output.
    # The normal SwitchingOrder file keeps the standard port-type mapping.
    if "Eagle Eye - Fund Source" in output.columns:
        output["FdSrc"] = output["Eagle Eye - Fund Source"]

    output["Eagle Eye Match Status"] = output["AccountNo"].map(
        lambda acct: "Matched" if acct in eagle_lookup else "Missing"
    )

    return output


# ------------------------------------------------------------
# ALLOCATION HELPERS
# ------------------------------------------------------------
def load_allocation(xls, sheet):
    try:
        df = pd.read_excel(xls, sheet_name=sheet).rename(
            columns=lambda x: str(x).strip().lower()
        )
    except Exception:
        return None

    if "mgr_code" not in df.columns or "fund_code" not in df.columns:
        return None

    if "weights" not in df.columns and "weights (%)" in df.columns:
        df.rename(columns={"weights (%)": "weights"}, inplace=True)
    if "weights" not in df.columns:
        return None

    df["mgr_code"] = df["mgr_code"].apply(
        lambda x: str(int(float(x))).zfill(3)
        if pd.notna(x) and str(x).strip() not in ("", "nan")
        else ""
    )
    df["fund_code"] = df["fund_code"].apply(
        lambda x: str(int(float(x))).zfill(3)
        if pd.notna(x) and str(x).strip() not in ("", "nan")
        else ""
    )
    df["weights"] = pd.to_numeric(df["weights"], errors="coerce").fillna(0)

    if not df.empty and df["weights"].max() > 1:
        df["weights"] = df["weights"] / 100

    df["key"] = df["mgr_code"] + df["fund_code"]
    return df


def prepare_smart_tab(xls, sheet):
    df = pd.read_excel(xls, sheet_name=sheet, dtype={"fund_code": str})
    df.columns = [str(c).strip() for c in df.columns]

    if "fund_code" in df.columns:
        df["fund_code"] = df["fund_code"].apply(
            lambda x: str(int(float(x))).zfill(3) if pd.notna(x) else x
        )

    weight_col = "weights (%)" if "weights (%)" in df.columns else "weights"
    if weight_col in df.columns:
        df[weight_col] = pd.to_numeric(df[weight_col], errors="coerce")
        non_null = df[weight_col].dropna()
        if not non_null.empty and non_null.le(1).all():
            df[weight_col] = (df[weight_col] * 100).round(2)
        if weight_col == "weights":
            df.rename(columns={"weights": "weights (%)"}, inplace=True)

    return df


# ------------------------------------------------------------
# FUND NAME MAP
# ------------------------------------------------------------
def build_fund_name_map(xls_list):
    fmap = {}

    for xls in xls_list:
        for sheet in xls.sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet).rename(
                    columns=lambda x: str(x).strip().lower()
                )
                if "mgr_code" not in df.columns or "fund_code" not in df.columns:
                    continue

                for _, row in df.iterrows():
                    try:
                        mgr = str(int(float(row["mgr_code"]))).zfill(3)
                        fund = str(int(float(row["fund_code"]))).zfill(3)
                        raw_name = row.get("fund name", "")
                        name = str(raw_name) if pd.notna(raw_name) else ""
                        if name:
                            fmap[(mgr, fund)] = name
                    except (ValueError, TypeError):
                        continue
            except Exception:
                continue

    return fmap


# ------------------------------------------------------------
# DELTA CALCULATION
# ------------------------------------------------------------
def compute_deltas(new_xls, old_xls, port_type):
    new_alloc = load_allocation(new_xls, port_type)
    old_alloc = load_allocation(old_xls, port_type)
    if new_alloc is None or old_alloc is None:
        return None

    new_cols = new_alloc[["key", "mgr_code", "fund_code", "weights"]].rename(
        columns={
            "weights": "new_w",
            "mgr_code": "mgr_new",
            "fund_code": "fund_new",
        }
    )
    old_cols = old_alloc[["key", "mgr_code", "fund_code", "weights"]].rename(
        columns={
            "weights": "old_w",
            "mgr_code": "mgr_old",
            "fund_code": "fund_old",
        }
    )

    merged = pd.merge(new_cols, old_cols, on="key", how="outer")
    merged["new_w"] = pd.to_numeric(merged["new_w"], errors="coerce").fillna(0)
    merged["old_w"] = pd.to_numeric(merged["old_w"], errors="coerce").fillna(0)
    merged["mgr_code"] = merged["mgr_new"].fillna(merged["mgr_old"])
    merged["fund_code"] = merged["fund_new"].fillna(merged["fund_old"])
    merged["delta"] = (merged["new_w"] - merged["old_w"]).round(6)

    return merged[merged["delta"] != 0].copy()


# ------------------------------------------------------------
# CORE TRADE CALCULATION
# ------------------------------------------------------------
def compute_switch_trades(rebal_df, new_xls, old_xls, fund_name_map):
    """
    Generate switch rows using the current many-to-many logic:
    every switch-out fund is split across all switch-in funds according to
    relative positive allocation deltas.
    """
    rebal_df = rebal_df.copy()
    rebal_df.columns = [str(c).strip() for c in rebal_df.columns]

    required_columns = {"svc_acct", "port_type", "code", "qty", "mkt_val", "total_pf"}
    missing = required_columns - set(rebal_df.columns)
    if missing:
        raise ValueError(
            "Rebalance input is missing required column(s): "
            + ", ".join(sorted(missing))
        )

    rebal_df["svc_acct"] = normalise_account_number(rebal_df["svc_acct"])
    rebal_df["port_type"] = (
        rebal_df["port_type"]
        .astype(str)
        .str.strip()
        .str.replace(r"^T_", "", regex=True)
    )
    rebal_df["code"] = rebal_df["code"].apply(
        lambda x: str(int(float(x))).zfill(6)
        if pd.notna(x) and str(x).strip() not in ("", "nan")
        else ""
    )
    rebal_df["MgrCd"] = rebal_df["code"].str[:3]
    rebal_df["FundCd"] = rebal_df["code"].str[-3:]
    rebal_df["qty"] = pd.to_numeric(rebal_df["qty"], errors="coerce").fillna(0)
    rebal_df["mkt_val"] = pd.to_numeric(rebal_df["mkt_val"], errors="coerce").fillna(0)
    rebal_df["total_pf"] = pd.to_numeric(rebal_df["total_pf"], errors="coerce").fillna(0)

    rebal_lookup = {}
    for _, row in rebal_df.iterrows():
        rebal_lookup[(row["svc_acct"], row["MgrCd"], row["FundCd"])] = row

    today_str = dt.datetime.today().strftime("%Y%m%d")
    trade_rows = []
    valid_port_types = set(new_xls.sheet_names) & set(old_xls.sheet_names)

    for port_type in rebal_df["port_type"].dropna().unique():
        if port_type not in valid_port_types:
            continue

        deltas = compute_deltas(new_xls, old_xls, port_type)
        if deltas is None or deltas.empty:
            continue

        switch_out = deltas[deltas["delta"] < 0].copy()
        switch_in = deltas[deltas["delta"] > 0].copy()
        if switch_out.empty or switch_in.empty:
            continue

        total_positive_delta = switch_in["delta"].sum()
        total_negative_delta = abs(switch_out["delta"].sum())

        if round(total_positive_delta, 6) != round(total_negative_delta, 6):
            print(
                f"WARNING [{port_type}]: positive delta "
                f"{total_positive_delta:.6f} does not equal negative delta "
                f"{total_negative_delta:.6f}"
            )

        if total_positive_delta <= 0:
            continue

        in_list = switch_in.to_dict("records")
        n_in = len(in_list)

        # Standard output uses the existing portfolio-type mapping.
        fdsrc = PORT_TYPE_FUND_SOURCE.get(
            port_type,
            "SRS-IA" if port_type.endswith("_SRS") else "Cash",
        )
        divopt = "Withdraw" if port_type.startswith("MD") else "Reinvest"

        old_alloc = load_allocation(old_xls, port_type)
        new_alloc = load_allocation(new_xls, port_type)
        old_w_map = {
            row["mgr_code"] + row["fund_code"]: row["weights"]
            for _, row in old_alloc.iterrows()
        }
        new_w_map = {
            row["mgr_code"] + row["fund_code"]: row["weights"]
            for _, row in new_alloc.iterrows()
        }

        unique_accts = rebal_df.loc[
            rebal_df["port_type"] == port_type, "svc_acct"
        ].unique()

        for acct_no in unique_accts:
            for _, out_row in switch_out.iterrows():
                mgr_out = out_row["mgr_code"]
                fund_out = out_row["fund_code"]
                key = mgr_out + fund_out

                rebal_row = rebal_lookup.get((acct_no, mgr_out, fund_out))
                if rebal_row is None:
                    continue

                qty = rebal_row["qty"]
                mkt_val = rebal_row["mkt_val"]
                total_pf = rebal_row["total_pf"]
                if qty <= 0:
                    continue

                old_w = old_w_map.get(key, 0)
                new_w = new_w_map.get(key, 0)

                if new_w == 0:
                    units_out = qty
                elif old_w > 0:
                    units_out = round(((old_w - new_w) / old_w) * qty, 3)
                else:
                    continue

                if units_out <= 0:
                    continue

                units_before = round(qty, 3)
                total_units_out = round(units_out, 3)
                units_remaining = round(qty - units_out, 3)
                mv_before = round(mkt_val, 2)
                account_aum = round(total_pf, 2)
                post_weight = (
                    round(
                        ((units_remaining / units_before) * mv_before) / account_aum * 100,
                        2,
                    )
                    if units_before > 0 and account_aum > 0
                    else 0.0
                )
                target_weight = round(new_w * 100, 2)
                diff = round(post_weight - target_weight, 2)
                so_name = fund_name_map.get((mgr_out, fund_out), "")

                allocated = 0.0
                for index, in_row in enumerate(in_list):
                    is_last = index == n_in - 1
                    mgr_in = in_row["mgr_code"]
                    fund_in = in_row["fund_code"]
                    delta_in = in_row["delta"]
                    si_name = fund_name_map.get((mgr_in, fund_in), "")

                    if is_last:
                        units_this = round(units_out - allocated, 3)
                    else:
                        units_this = round(
                            units_out * (delta_in / total_positive_delta),
                            3,
                        )
                        allocated += units_this

                    if units_this <= 0:
                        continue

                    order_type = "SO" if mgr_out == mgr_in else "ESO"

                    trade_rows.append(
                        {
                            "AccountNo": acct_no,
                            "OrderType": order_type,
                            "MgrCd": mgr_out,
                            "FundCd": fund_out,
                            "FdSrc": fdsrc,
                            "SIMgrCd": mgr_in,
                            "SIFundCd": fund_in,
                            "NetSalesChg": 0,
                            "SwChg": 0,
                            "DivOpt": divopt,
                            "Units": units_this,
                            "FACode": "",
                            "TransDate": today_str,
                            "Portfolio Type": port_type,
                            "Switch Out Fund Name": so_name,
                            "Switch In Fund Name": si_name,
                            "SO: Units before Rebalance": units_before,
                            "Total Units SO": total_units_out,
                            "SO: Units Remaining after Rebalance": units_remaining,
                            "SO: MV of Fund before Rebalance (SGD)": mv_before,
                            "Account Portfolio Amount (SGD)": account_aum,
                            "SO Fund Weight (%) after Rebalance": post_weight,
                            "Target Allocation of Switch Out Fund (%)": target_weight,
                            "Diff": diff,
                        }
                    )

    return pd.DataFrame(trade_rows) if trade_rows else pd.DataFrame()


# ------------------------------------------------------------
# EXCEL FORMATTING
# ------------------------------------------------------------
def fmt_sheet(ws, header_fill, freeze="A2", header_row=1, data_start_row=2):
    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = BORDER

    ws.row_dimensions[header_row].height = 30

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start_row),
        start=data_start_row,
    ):
        fill = ALT_FILL_LIGHT if row_idx % 2 == 0 else ALT_FILL_NONE
        for cell in row:
            cell.font = BODY_FONT
            cell.border = BORDER
            if cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                cell.fill = fill

    if freeze:
        ws.freeze_panes = freeze

    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        first = next((cell for cell in col if hasattr(cell, "column_letter")), None)
        if first:
            ws.column_dimensions[first.column_letter].width = min(
                max(max_len + 2, 10),
                50,
            )


# ------------------------------------------------------------
# STANDARD OUTPUT WRITER
# ------------------------------------------------------------
def write_switch_file(path, trades_df, new_xls, relevant_sheets):
    trade_cols = [
        "AccountNo",
        "OrderType",
        "MgrCd",
        "FundCd",
        "FdSrc",
        "SIMgrCd",
        "SIFundCd",
        "NetSalesChg",
        "SwChg",
        "DivOpt",
        "Units",
        "FACode",
        "TransDate",
    ]
    check_cols = trade_cols + [
        "Portfolio Type",
        "Switch Out Fund Name",
        "Switch In Fund Name",
        "SO: Units before Rebalance",
        "Total Units SO",
        "SO: Units Remaining after Rebalance",
        "SO: MV of Fund before Rebalance (SGD)",
        "Account Portfolio Amount (SGD)",
        "SO Fund Weight (%) after Rebalance",
        "Target Allocation of Switch Out Fund (%)",
        "Diff",
    ]

    trades_out = trades_df[trade_cols].copy()
    check_out = trades_df[check_cols].copy()

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        trades_out.to_excel(writer, sheet_name="Trades", index=False)

        ws_check = writer.book.create_sheet("Check")
        writer.sheets["Check"] = ws_check
        ws_check.cell(
            row=1,
            column=1,
            value=(
                "CHECK TAB - Green = Diff within +/-1% of target. "
                "Yellow = deviation above 1% and requires review."
            ),
        )

        for col_idx, header in enumerate(check_cols, start=1):
            ws_check.cell(row=2, column=col_idx, value=header)

        for row_idx, row_data in enumerate(
            check_out.itertuples(index=False),
            start=3,
        ):
            for col_idx, value in enumerate(row_data, start=1):
                ws_check.cell(row=row_idx, column=col_idx, value=value)

        writer.book.create_sheet("New Portfolio Allocation >>")

        for sheet in relevant_sheets:
            if sheet in new_xls.sheet_names:
                prepare_smart_tab(new_xls, sheet).to_excel(
                    writer,
                    sheet_name=sheet,
                    index=False,
                )

    wb = load_workbook(path)
    fmt_sheet(wb["Trades"], HEADER_FILL_TRADE)

    ws_check = wb["Check"]
    note_cell = ws_check.cell(row=1, column=1)
    note_cell.font = Font(
        bold=True,
        italic=True,
        name="Arial",
        size=10,
        color="375623",
    )
    note_cell.alignment = Alignment(horizontal="left")
    ws_check.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(check_cols),
    )
    fmt_sheet(
        ws_check,
        HEADER_FILL_CHECK,
        freeze="A3",
        header_row=2,
        data_start_row=3,
    )

    diff_col = get_column_letter(check_cols.index("Diff") + 1)
    for row_no in range(3, len(check_out) + 3):
        cell = ws_check[f"{diff_col}{row_no}"]
        try:
            value = float(cell.value) if cell.value is not None else None
            if value is not None:
                cell.fill = OK_FILL if abs(value) <= 1.0 else WARN_FILL
        except (TypeError, ValueError):
            pass

    for sheet in relevant_sheets:
        if sheet in wb.sheetnames:
            fmt_sheet(wb[sheet], HEADER_FILL_ALLOC)

    desired = ["Trades", "Check", "New Portfolio Allocation >>"] + relevant_sheets
    wb._sheets.sort(
        key=lambda sheet: desired.index(sheet.title)
        if sheet.title in desired
        else 999
    )

    wb.save(path)
    print(f"  Saved standard switch file: {path}")


# ------------------------------------------------------------
# EAGLE EYE OUTPUT WRITER - NEW
# ------------------------------------------------------------
def write_eagle_eye_file(path, eagle_output_df):
    """Write the separate Eagle Eye-enriched workbook."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        eagle_output_df.to_excel(writer, sheet_name="Eagle Eye Trades", index=False)

        missing = eagle_output_df.loc[
            eagle_output_df["Eagle Eye Match Status"] == "Missing"
        ].copy()
        if not missing.empty:
            missing.to_excel(writer, sheet_name="Missing Accounts", index=False)

    wb = load_workbook(path)
    fmt_sheet(wb["Eagle Eye Trades"], HEADER_FILL_EAGLE)

    if "Missing Accounts" in wb.sheetnames:
        fmt_sheet(wb["Missing Accounts"], HEADER_FILL_CHECK)

    wb.save(path)
    print(f"  Saved Eagle Eye file: {path}")


# ------------------------------------------------------------
# MAIN
# ------------------------------------------------------------
def main():
    print("Switch Order Generator - initialising...")
    time.sleep(0.5)

    show_msg("Step 1", "Select the REBALANCE INPUT file.")
    rebalance_path = prompt_file(
        "Step 1: Rebalance Input File",
        filetypes=[
            ("Excel files", "*.xlsx *.xls *.xlsm"),
            ("All files", "*.*"),
        ],
    )
    if not rebalance_path:
        sys.exit()

    show_msg("Step 2", "Select the NEW ALLOCATION file.")
    new_alloc_path = prompt_file(
        "Step 2: New Allocation File",
        filetypes=[
            ("Excel files", "*.xlsx *.xls *.xlsm"),
            ("All files", "*.*"),
        ],
    )
    if not new_alloc_path:
        sys.exit()

    show_msg("Step 3", "Select the OLD ALLOCATION file.")
    old_alloc_path = prompt_file(
        "Step 3: Old Allocation File",
        filetypes=[
            ("Excel files", "*.xlsx *.xls *.xlsm"),
            ("All files", "*.*"),
        ],
    )
    if not old_alloc_path:
        sys.exit()

    # NEW: Eagle Eye is optional and creates a separate output workbook.
    generate_eagle_eye = ask_yes_no(
        "Eagle Eye Output",
        "Do you want to generate a separate Eagle Eye-enriched output file?",
    )

    eagle_df = None
    if generate_eagle_eye:
        show_msg("Step 4", "Select the Eagle Eye account export file.")
        eagle_path = prompt_file(
            "Step 4: Eagle Eye Account Export",
            filetypes=[
                ("Excel files", "*.xlsx *.xls *.xlsm"),
                ("All files", "*.*"),
            ],
        )
        if not eagle_path:
            show_msg(
                "Cancelled",
                "The Eagle Eye export is required to create the Eagle Eye output.",
            )
            sys.exit()
        eagle_df = clean_df(load_excel_safe(eagle_path))

    print("Loading files...")
    rebal_df = clean_df(load_excel_safe(rebalance_path))
    new_xls = pd.ExcelFile(new_alloc_path)
    old_xls = pd.ExcelFile(old_alloc_path)

    rebal_df["svc_acct"] = normalise_account_number(rebal_df["svc_acct"])
    rebal_df["port_type"] = (
        rebal_df["port_type"]
        .astype(str)
        .str.strip()
        .str.replace(r"^T_", "", regex=True)
    )

    rebal_port_types = set(rebal_df["port_type"].unique())
    common_port_types = (
        rebal_port_types
        & set(new_xls.sheet_names)
        & set(old_xls.sheet_names)
    )

    changed_port_types = []
    for port_type in sorted(common_port_types):
        deltas = compute_deltas(new_xls, old_xls, port_type)
        if deltas is not None and not deltas.empty:
            changed_port_types.append(port_type)

    if not changed_port_types:
        show_msg(
            "No Changes",
            "No allocation changes detected between the old and new files.",
        )
        sys.exit()

    print(f"  Portfolio types with changes: {changed_port_types}")

    print("Building fund name map...")
    fund_name_map = build_fund_name_map([new_xls, old_xls])

    print("Computing switch trades...")
    trades_df = compute_switch_trades(
        rebal_df,
        new_xls,
        old_xls,
        fund_name_map,
    )

    if trades_df.empty:
        show_msg(
            "No Trades",
            "No trades were generated. Check the input files and allocation changes.",
        )
        sys.exit()

    print(f"  {len(trades_df)} trade rows generated.")
    pairs = trades_df.groupby(
        ["Portfolio Type", "MgrCd", "FundCd", "SIMgrCd", "SIFundCd"]
    ).size()
    print("\nSwitch pairs generated:")
    print(pairs.to_string())

    relevant_sheets = [
        sheet
        for sheet in new_xls.sheet_names
        if sheet in trades_df["Portfolio Type"].unique()
    ]

    show_msg("Output Folder", "Select a folder for the generated output file(s).")
    output_folder = prompt_folder("Select Output Folder")
    if not output_folder:
        sys.exit()

    today_str = dt.datetime.today().strftime("%Y%m%d")
    standard_output_path = os.path.join(
        output_folder,
        f"SwitchingOrder_{today_str}.xlsx",
    )

    print("\nWriting standard switch-order file...")
    write_switch_file(
        path=standard_output_path,
        trades_df=trades_df,
        new_xls=new_xls,
        relevant_sheets=relevant_sheets,
    )

    # NEW: Write a second file only when Eagle Eye was selected.
    eagle_output_path = None
    if generate_eagle_eye:
        print("Building Eagle Eye account lookup...")
        eagle_lookup = build_eagle_eye_lookup(eagle_df)
        eagle_output_df = create_eagle_eye_output(trades_df, eagle_lookup)

        missing_count = int(
            (eagle_output_df["Eagle Eye Match Status"] == "Missing").sum()
        )
        if missing_count:
            print(
                f"WARNING: {missing_count} trade row(s) could not be matched "
                "to an Eagle Eye service account."
            )

        eagle_output_path = os.path.join(
            output_folder,
            f"EagleEye_SwitchingOrder_{today_str}.xlsx",
        )
        print("Writing Eagle Eye output file...")
        write_eagle_eye_file(eagle_output_path, eagle_output_df)

    final_message = f"Standard output:\n{standard_output_path}"
    if eagle_output_path:
        final_message += f"\n\nEagle Eye output:\n{eagle_output_path}"

    show_msg("Completed", final_message)
    print(f"\nDone.\n{final_message}")


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        show_msg("Error", str(exc))
        print(f"ERROR: {exc}")
        sys.exit(1)
